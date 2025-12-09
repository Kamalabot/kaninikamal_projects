
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

# Step 1: Load or create budget data
budget_data = {
    'Department': ['Finance', 'HR', 'Sales', 'Operations', 'Marketing'],
    'Budget Amount': [10000, 8000, 15000, 12000, 7000],
    'Actual Spend': [9500, 8200, 16500, 11000, 6800]
}

df = pd.DataFrame(budget_data)

# Step 2: Calculate variance and identify over-budget depts
df['Variance'] = df['Budget Amount'] - df['Actual Spend']
df['Status'] = df['Variance'].apply(lambda x: 'Over Budget' if x < 0 else 'Within Budget')

# Step 3: Generate Excel report with conditional formatting
output_file = f'Budget_Report_{datetime.now().strftime("%Y%m%d")}.xlsx'

with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='Details', index=False)

    # Add summary pivot
    summary = df.groupby('Status')['Variance'].sum().to_frame()
    summary.to_excel(writer, sheet_name='Summary')

    # Formatting
    workbook = writer.book
    sheet = writer.sheets['Details']

    red_fill = PatternFill(start_color='FF0000', fill_type='solid')
    for row in sheet.iter_rows(min_row=2, max_col=5):
        if row[4].value == 'Over Budget':
            for cell in row:
                cell.fill = red_fill

print(f"✓ Budget report saved: {output_file}")
print(f"\nOver-Budget Alert:")
over_budget = df[df['Variance'] < 0]
for _, row in over_budget.iterrows():
    print(f"  - {row['Department']}: ${abs(row['Variance']):,.0f} over")

# Step 4 (Optional): Email report
# msg = MIMEMultipart()
# msg['Subject'] = f'Monthly Budget Report - {datetime.now().strftime("%B %Y")}'
# ... attach output_file and send via SMTP
